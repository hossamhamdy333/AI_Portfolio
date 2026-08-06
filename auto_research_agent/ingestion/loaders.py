"""
Multimodal loaders — all free/local, no paid APIs.

- PDF / DOCX / XLSX  -> text via pypdf / python-docx / openpyxl
- Audio / video       -> text via faster-whisper (local, free, replaces paid Whisper API)
- Images / graphs     -> text via Groq's free vision model (llama-3.2-11b-vision-preview)
                         with pytesseract OCR as an offline fallback if no API key is set.
"""
import os
from typing import List, Dict

from pypdf import PdfReader
import docx
import openpyxl


def load_pdf(path: str) -> List[Dict]:
    reader = PdfReader(path)
    out = []
    for i, page in enumerate(reader.pages):
        text = page.extract_text() or ""
        if text.strip():
            out.append({"text": text, "source": os.path.basename(path), "page": i + 1, "modality": "text"})
    return out


def load_docx(path: str) -> List[Dict]:
    d = docx.Document(path)
    text = "\n".join(p.text for p in d.paragraphs if p.text.strip())
    return [{"text": text, "source": os.path.basename(path), "page": None, "modality": "text"}]


def load_xlsx(path: str) -> List[Dict]:
    wb = openpyxl.load_workbook(path, data_only=True)
    out = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        rows = []
        for row in ws.iter_rows(values_only=True):
            if any(cell is not None for cell in row):
                rows.append(", ".join(str(c) for c in row if c is not None))
        if rows:
            out.append({
                "text": f"Sheet: {sheet}\n" + "\n".join(rows),
                "source": os.path.basename(path),
                "page": sheet,
                "modality": "table",
            })
    return out


_whisper_model = None


def _get_whisper():
    global _whisper_model
    if _whisper_model is None:
        from faster_whisper import WhisperModel
        # "base" runs fine on Colab CPU/GPU and is free — no API call.
        _whisper_model = WhisperModel("base", device="auto", compute_type="int8")
    return _whisper_model


def transcribe_audio_video(path: str) -> List[Dict]:
    """Transcribes local audio/video files. For YouTube links, download audio first with yt-dlp."""
    model = _get_whisper()
    segments, _ = model.transcribe(path)
    text = " ".join(seg.text for seg in segments)
    return [{"text": text, "source": os.path.basename(path), "page": None, "modality": "audio"}]


def download_youtube_audio(url: str, out_dir: str = "/content/yt_audio") -> str:
    """Free audio download via yt-dlp, no API key. Returns local mp3 path."""
    import yt_dlp
    os.makedirs(out_dir, exist_ok=True)
    ydl_opts = {
        "format": "bestaudio/best",
        "outtmpl": os.path.join(out_dir, "%(id)s.%(ext)s"),
        "postprocessors": [{"key": "FFmpegExtractAudio", "preferredcodec": "mp3"}],
        "quiet": True,
    }
    with yt_dlp.YoutubeDL(ydl_opts) as ydl:
        info = ydl.extract_info(url, download=True)
        return os.path.join(out_dir, f"{info['id']}.mp3")


def describe_image(path: str, groq_client=None) -> List[Dict]:
    """
    Summarizes an image/graph into text.
    Uses Groq's free vision model if a client is passed; otherwise falls back to
    local OCR (pytesseract) which reads text in the image but can't describe charts.
    """
    if groq_client is not None:
        import base64
        with open(path, "rb") as f:
            b64 = base64.b64encode(f.read()).decode()
        resp = groq_client.chat.completions.create(
            model="llama-3.2-11b-vision-preview",
            messages=[{
                "role": "user",
                "content": [
                    {"type": "text", "text": "Describe this image/graph factually, including any numbers, "
                                              "trends, or labels visible."},
                    {"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{b64}"}},
                ],
            }],
        )
        text = resp.choices[0].message.content
    else:
        import pytesseract
        from PIL import Image
        text = pytesseract.image_to_string(Image.open(path))
    return [{"text": text, "source": os.path.basename(path), "page": None, "modality": "image"}]


LOADERS = {
    ".pdf": load_pdf,
    ".docx": load_docx,
    ".xlsx": load_xlsx,
    ".mp3": transcribe_audio_video,
    ".mp4": transcribe_audio_video,
    ".wav": transcribe_audio_video,
    ".png": describe_image,
    ".jpg": describe_image,
    ".jpeg": describe_image,
}


def load_any(path: str, groq_client=None) -> List[Dict]:
    ext = os.path.splitext(path)[1].lower()
    fn = LOADERS.get(ext)
    if fn is None:
        raise ValueError(f"No loader for {ext}")
    if fn is describe_image:
        return fn(path, groq_client=groq_client)
    return fn(path)
